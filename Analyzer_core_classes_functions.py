# Main Classes, functions, APIs

import mysql.connector
import zillow
import datetime
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
import locale
import openpyxl
import pprint
import logging
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver


class HometownLocator(object):
    def __init__(self, street, state, city, short_state, xls_name):
        self.driver = webdriver.Chrome('/Users/alexdezho/Documents/Personal/chromedriver')
        # class params for multiple use with multiple functions
        self.stateformetro = state
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.replace(" ", "").lower() + " " + self.short_state.lower()
        # urls
        self.googleMaps_url = "https://www.google.com/maps/"
        self.coordinate_search_url = "https://" + state.replace(" ", "").lower() + ".hometownlocator.com/maps"
        self.metropol_Tool_url = "https://www.huduser.gov/portal/datasets/geotool/select_Geography.odn"
        self.metropolitan_url = "https://" + str(state).replace(" ", "").lower() + ".hometownlocator.com/cities/msa/"
        self.coordinates_url_track = ''
        self.census_block_url = ''
        self.census_track_url = ''
        self.county_url = " "
        self.city_url = ''
        self.zip_code_url = ' '
        self.metropolitan_url_htl = ' '
        self.google_maps_link = ' '
        # params for DB
        self.zip_code = ''
        self.coord = " "
        self.county = " "
        self.state_search = " "
        self.county_search = " "
        self.metropolitan_name = " "
        self.index = " "
        # dictionaries
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'county': '',
            'zip_code': '',
            'metropolitan': '',
            'link_google_maps': '',
            'coordinates': ''
        }
        self.dict_block = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_track = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_zip_code = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_city = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_county = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_metro = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def google_Maps_Addr_Coord(self):
        try:
            driver = self.driver
            driver.get(self.googleMaps_url)
            # await command waiting till we find the element then continue
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="searchboxinput"]')))
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(self.full_addr)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="searchbox-searchbutton"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom1
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom2
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom3
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom4
            time.sleep(3)
            self.coord = self.driver.current_url
            index = self.coord.find("@") + 1
            self.coord = self.coord[index:index + 21]
            self.coord = self.coord.replace(",", ":")
            self.dict_basic_info['coordinates'] = self.coord
            self.google_maps_link = driver.current_url
            self.dict_basic_info['link_google_maps'] = self.google_maps_link
            print("google coordinates is: {}".format(self.coord))
            print('using url {}'.format(self.coordinate_search_url))
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('1 - scrolling to enter address')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            print('preparing to locate general info')
            print('starting with trying to locate zip code')
            self.zip_code_url = driver.find_element_by_partial_link_text('ZIP Code ').text
            index1 = self.zip_code_url.find('Code') + 5
            self.zip_code_url = self.zip_code_url[index1:]
            self.dict_basic_info['zip_code'] = self.zip_code_url
            print('zip code located')
            print(self.dict_basic_info['zip_code'])
            self.zip_code_url = 'https://' + self.state.replace(" ", "").lower() + '.hometownlocator.com/zip-codes/data,zipcode,' + self.zip_code_url + '.cfm'
            print('trying to locate county url')
            self.county = driver.find_element_by_partial_link_text('County').text
            self.dict_basic_info['county'] = self.county
            self.county_url = "https://" + self.state.replace(" ", "").lower() + ".hometownlocator.com/" + self.short_state.lower() + "/" + str(self.county)[:-7].lower() + "/"
            print('county is {}'.format(self.county))
            print('county url located')
            print(self.county_url)
            print('trying to locate city url')
            self.city_url = "https://" + self.state.replace(" ", "").lower() + ".hometownlocator.com/" + self.short_state.lower() + "/" + str(self.county)[:-7].lower() + "/" + self.city.lower() + ".cfm"
            print('city url detected')
            print(self.city_url)
        except:
            print('failed - to locate all basic information')
            return False

    def metropolitan_area_Look_Up_Tool(self):
        # getting metropolitan name by using state name and county name
        self.state_search = self.state + " - " + self.short_state.upper()
        self.county_search = self.county + ", " + self.short_state.upper()
        try:
            driver = self.driver
            driver.get(self.metropol_Tool_url)
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="select_geo_options"]/form[1]/div[1]/select')))
            # select state - table 1
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/div[1]/select')
            Select(driver.find_element_by_tag_name('select')).select_by_visible_text(self.state_search)
            time.sleep(1)
            logging.debug('first table state selected,success')
            # select - table 2
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/div[2]/select')
            Select(driver.find_element_by_id("countyselect")).select_by_visible_text(self.county_search)
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/input').click()
            # locate metropolitan name
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/p[2]/em')))
            self.metropolitan_name = driver.find_element_by_xpath('/html/body/p[2]/em').text
            self.metropolitan_name = str(self.metropolitan_name)
            self.index = self.metropolitan_name.find('-')
            self.metropolitan_name = self.metropolitan_name[:self.index]
            self.dict_basic_info['metropolitan'] = self.metropolitan_name
            logging.debug('Metropolitan name: {}'.format(self.metropolitan_name))
            print('Metropolitan found name: {}'.format(self.metropolitan_name))
        except:
            print('Metropolitan name Failed to found')
            return False

    def metro_to_url(self):
        try:
            driver = self.driver
            driver.get(self.metropolitan_url)
            time.sleep(1)
            driver.execute_script("window.scrollTo(0,615)")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, self.metropolitan_name)))
            driver.find_element_by_partial_link_text(self.metropolitan_name).click()
            time.sleep(1)
            self.metropolitan_url_htl = driver.current_url
            #print('metropolitan url located!')
            logging.debug("Metro url: {}".format(self.metropolitan_url_htl))
            print("Metro url: {}".format(self.metropolitan_url_htl))
            return self.metropolitan_url_htl
        except:
            print('failed to locate metro url')
            return False

    def return_block_url(self):
        try:
            driver = self.driver
            print('trying to locate block url')
            print(self.coordinate_search_url)
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('scrolling to search address')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(3)
            self.census_block_url = driver.find_element_by_partial_link_text('(Census Block Group)').click()
            time.sleep(5)
            self.census_block_url = driver.current_url
            print(driver.current_url)
        except:
            print('failed to locate block url - not from 404 error')

        return self.census_block_url

    def return_track_url(self):
        try:
            driver = self.driver
            print('trying to locate track url')
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('1')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(3)
            self.census_track_url = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[3]/ul/li[2]/a').click()
            time.sleep(5)
            self.census_track_url = driver.current_url
            a = ""
            try:
                a = driver.find_element_by_xpath('//*[@id="content"]/div/fieldset/h2').text
            except:
                print('url should be founded')

            if a == '404 - File or directory not found.':
                print('failed to locate track url')
                self.dict_track["Total_Population"] = 'NA'
                self.dict_track["Population_Growth_2010_2019"] = 'NA'
                self.dict_track["Population_Growth_2019_2024"] = 'NA'
                self.dict_track["Median_Household_Income"] = 'NA'
                self.dict_track["Average_Household_Income"] = 'NA'
                self.dict_track["Total_Housing_Units"] = 'NA'
                self.dict_track["Owner_Occupied_HU"] = 'NA'
                self.dict_track["Renter_Occupied_HU"] = 'NA'
                self.dict_track["Vacant_Housing_Units"] = 'NA'
                self.dict_track["Median_Home_Value"] = 'NA'
            else:
                print('track url located')
                print(self.census_track_url)
        except:
            print('failed to locate track url not from 404 error')
        return self.census_track_url

    def return_zip_code_url(self):
        try:
            driver = self.driver
            print('trying to locate zip code url')
            #print(self.coordinate_search_url)
            driver.get(self.coordinate_search_url)
            time.sleep(3)
            driver.execute_script("window.scrollTo(0,615)")
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[3]/ul/li[4]/a')))
            self.zip_code_url = driver.find_element_by_partial_link_text('ZIP Code ').click()
            time.sleep(3)
            self.zip_code_url = driver.current_url
            a = ""
            try:
                a = driver.find_element_by_xpath('//*[@id="content"]/div/fieldset/h2').text
            except:
                print('url should be founded')

            if a == '404 - File or directory not found.':
                print('failed to locate zip code url')
                self.dict_zip_code["Total_Population"] = 'NA'
                self.dict_zip_code["Population_Growth_2010_2019"] = 'NA'
                self.dict_zip_code["Population_Growth_2019_2024"] = 'NA'
                self.dict_zip_code["Median_Household_Income"] = 'NA'
                self.dict_zip_code["Average_Household_Income"] = 'NA'
                self.dict_zip_code["Total_Housing_Units"] = 'NA'
                self.dict_zip_code["Owner_Occupied_HU"] = 'NA'
                self.dict_zip_code["Renter_Occupied_HU"] = 'NA'
                self.dict_zip_code["Vacant_Housing_Units"] = 'NA'
                self.dict_zip_code["Median_Home_Value"] = 'NA'
            else:
                print('zip code url located')
                print(self.zip_code_url)

        except:

            print('failed to locate zip code url not from 404 error')

        return self.zip_code_url

    def return_county_url(self):
        try:
            print('trying to locate county url')
            driver = self.driver
            driver.get(self.coordinate_search_url)
            time.sleep(20)
            print('1')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(5)
            driver.find_element_by_partial_link_text('County').click()
            time.sleep(5)
            self.county_url = driver.current_url
            return self.county_url
        except:
            print('could not return county url')

    def return_city_url(self):
        return self.city_url

    def return_metro_url(self):
        return self.metropolitan_url_htl

    def HTML_to_dictionary(self, url):
        dict = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        try:
            driver = self.driver
            driver.get(url)
            print(url)
            print('trying to locate elements path 1')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]')))
            print('printing for checking the first element')
            print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text)
            time.sleep(2)
            print('1')
            dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text
            print('2')
            dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[9]/table/tbody/tr[2]/td[2]').text
            print('3')
            dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[9]/table/tbody/tr[2]/td[3]').text
            print('4')
            dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[11]/td[2]').text
            print('5')
            dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[12]/td[2]').text
            print('6')
            dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text
            print('7')
            dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[3]/td[2]').text
            print('8')
            dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[4]/td[2]').text
            print('9')
            dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[5]/td[2]').text
            print('10')
            dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[6]/td[2]').text
            print('11')
            dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[11]/td[2]').text
            print('12')
            dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[12]/td[2]').text
            print('13')
            dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[13]/td[2]').text
            print('14')
            print('HTML page params was copied to dict success------------------------------')
        except:
            try:
                driver = self.driver
                print('trying to locate elements with another path 2')
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]')))
                print('printing for checking the first element')
                print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]').text)
                time.sleep(2)
                print('1')
                dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]').text
                print('2')
                dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[7]/table/tbody/tr[2]/td[2]').text
                print('3')
                dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[7]/table/tbody/tr[2]/td[3]').text
                print('4')
                dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[11]/td[2]').text
                print('5')
                dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[12]/td[2]').text
                print('6')
                dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text
                print('7')
                dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[3]/td[2]').text
                print('8')
                dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[4]/td[2]').text
                print('9')
                dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[5]/td[2]').text
                print('10')
                dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[6]/td[2]').text
                print('11')
                dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[11]/td[2]').text
                print('12')
                dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[12]/td[2]').text
                print('13')
                dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[13]/td[2]').text
                print('14')
                print('HTML page params was copied to dict success------------------------------')
            except:
                try:
                    driver = self.driver
                    print('trying to locate elements with another path 3')
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]')))
                    print('printing for checking the first element')
                    print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text)
                    time.sleep(2)
                    print('1')
                    dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text
                    print('2')
                    dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[8]/table/tbody/tr[2]/td[2]').text
                    print('3')
                    dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[8]/table/tbody/tr[2]/td[3]').text
                    print('4')
                    dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[11]/td[2]').text
                    print('5')
                    dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[12]/td[2]').text
                    print('6')
                    dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text
                    print('7')
                    dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[3]/td[2]').text
                    print('8')
                    dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[4]/td[2]').text
                    print('9')
                    dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[5]/td[2]').text
                    print('10')
                    dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[6]/td[2]').text
                    print('11')
                    dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[11]/td[2]').text
                    print('12')
                    dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[12]/td[2]').text
                    print('13')
                    dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[13]/td[2]').text
                    print('14')
                    print('HTML page params was copied to dict success------------------------------')
                except:
                    try:
                        driver = self.driver
                        print('trying to locate elements with another third path 4')
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]')))
                        print('printing for checking the first element')
                        print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text)
                        time.sleep(2)
                        print('1')
                        dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text
                        print('2')
                        dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[10]/table/tbody/tr[2]/td[2]').text
                        print('3')
                        dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[10]/table/tbody/tr[2]/td[3]').text
                        print('4')
                        dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[11]/td[2]').text
                        print('5')
                        dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[12]/td[2]').text
                        print('6')
                        dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[2]/td[2]').text
                        print('7')
                        dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[3]/td[2]').text
                        print('8')
                        dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[4]/td[2]').text
                        print('9')
                        dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[5]/td[2]').text
                        print('10')
                        dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[6]/td[2]').text
                        print('11')
                        dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[11]/td[2]').text
                        print('12')
                        dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[12]/td[2]').text
                        print('13')
                        dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[13]/td[2]').text
                        print('14')
                        print('HTML page params was copied to dict success------------------------------')
                    except:
                        try:
                            driver = self.driver
                            print('trying to locate elements with another third path 5')
                            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]')))
                            print('printing for checking the first element')
                            print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]').text)

                            time.sleep(2)
                            print('1')
                            dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]').text
                            print('2')
                            dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[12]/table/tbody/tr[2]/td[2]').text
                            print('3')
                            dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[12]/table/tbody/tr[2]/td[3]').text
                            print('4')
                            dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[11]/td[2]').text
                            print('5')
                            dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[12]/td[2]').text
                            print('6')
                            dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[2]/td[2]').text
                            print('7')
                            dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[3]/td[2]').text
                            print('8')
                            dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[4]/td[2]').text
                            print('9')
                            dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[5]/td[2]').text
                            print('10')
                            dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[6]/td[2]').text
                            print('11')
                            dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[11]/td[2]').text
                            print('12')
                            dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[12]/td[2]').text
                            print('13')
                            dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[13]/td[2]').text
                            print('14')
                            print('HTML page params was copied to dict success------------------------------')
                        except:
                            print('elements not found')
                            dict["Total_Population"] = "NA"
                            dict["Population_Growth_2010_2019"] = 'NA'
                            dict["Population_Growth_2019_2024"] = 'NA'
                            dict["Median_Household_Income"] = 'NA'
                            dict["Average_Household_Income"] = 'NA'
                            dict["Total_Housing_Units"] = 'NA'
                            dict["Owner_Occupied_HU"] = 'NA'
                            dict["Renter_Occupied_HU"] = 'NA'
                            dict["Vacant_Housing_Units"] = 'NA'
                            dict["Median_Home_Value"] = 'NA'
                            dict["Total_Hoouseholds"] = 'NA'
                            dict["Avarage_Households_Size"] = 'NA'
                            dict["Family_Households"] = 'NA'
        return dict

    def params_to_dict_block(self, dict):
        try:
            self.dict_block = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_track(self, dict):
        try:
            self.dict_track = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_zip_code(self, dict):
        try:
            self.dict_zip_code = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_city(self, dict):
        try:
            self.dict_city = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_county(self, dict):
        try:
            self.dict_county = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_metro(self, dict):
        try:
            self.dict_metro = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    # print all dictionaries
    def printall(self):
        print('hi all dictionaries\n')
        pp = pprint.PrettyPrinter(indent=4)
        print('basic info: {}'.format(self.dict_basic_info))
        pp.pprint(self.dict_basic_info)
        print('block: {}'.format(self.dict_block))
        pp.pprint(self.dict_block)
        print('track: {}'.format(self.dict_track))
        pp.pprint(self.dict_track)
        print('zip_code: {}'.format(self.dict_zip_code))
        pp.pprint(self.dict_zip_code)
        print('city: {}'.format(self.dict_city))
        pp.pprint(self.dict_city)
        print('metro: {}'.format(self.dict_metro))
        pp.pprint(self.dict_metro)

    # copy all dictionaries to xls file
    def xls_new_sheet_for_search_create(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            if wb.sheetnames.count(self.full_addr[:25]) == 0:
                example_sheet = wb["example"]
                wb.copy_worksheet(example_sheet)
                # print(wb.sheetnames)
                new_sheet = wb['example Copy']
                new_sheet.title = self.full_addr[:25]
                # print(wb.sheetnames)
                wb.save(self.xls_name)
                print("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                wb.close()
                return True
            else:
                print("address was already searched & exists in database")
                wb.close()
                logging.debug("address was already searched & exists in database")
                return False
        except:
            print('faild to create xls file')

    def basic_Info_dict_to_xls(self):
        try:
            # opening xls
            print('opening XLS to save params')
            logging.debug('opening XLS to save params')
            print('the xls file name is: {}'.format(self.xls_name))
            wb = openpyxl.load_workbook(self.xls_name)
            print(wb.sheetnames)
            sheet = wb[self.full_addr[:25]]
            print(self.full_addr[:25])
            # google link
            sheet['B17'].value = self.google_maps_link
            sheet['B3'].value = self.dict_basic_info['zip_code']
            wb.save(self.xls_name)
            wb.close()
            return True
        except:
            print('failed to copy info to XLS ')
            logging.debug('failed to open XLS')
            return False

    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.full_addr[:25]]
            # print(wb.sheetnames)
            sheet['B7'].value = self.dict_block["Total_Population"]
            sheet['B8'].value = self.dict_block["Population_Growth_2010_2019"] + "(per year)"
            sheet['B9'].value = self.dict_block["Population_Growth_2019_2024"] + "(per year)"
            sheet['B10'].value = self.dict_block["Median_Household_Income"]
            sheet['B11'].value = self.dict_block["Average_Household_Income"]
            sheet['B12'].value = self.dict_block["Total_Housing_Units"]
            sheet['B13'].value = self.dict_block["Owner_Occupied_HU"]
            sheet['B14'].value = self.dict_block["Renter_Occupied_HU"]
            sheet['B15'].value = self.dict_block["Vacant_Housing_Units"]
            sheet['B16'].value = self.dict_block["Median_Home_Value"]
            sheet['B18'].value = self.dict_block["Total_Hoouseholds"]
            sheet['B19'].value = self.dict_block["Avarage_Households_Size"]
            sheet['B20'].value = self.dict_block["Family_Households"]

            sheet['C7'].value = self.dict_track["Total_Population"]
            sheet['C8'].value = self.dict_track["Population_Growth_2010_2019"] + "(per year)"
            sheet['C9'].value = self.dict_track["Population_Growth_2019_2024"] + "(per year)"
            sheet['C10'].value = self.dict_track["Median_Household_Income"]
            sheet['C11'].value = self.dict_track["Average_Household_Income"]
            sheet['C12'].value = self.dict_track["Total_Housing_Units"]
            sheet['C13'].value = self.dict_track["Owner_Occupied_HU"]
            sheet['C14'].value = self.dict_track["Renter_Occupied_HU"]
            sheet['C15'].value = self.dict_track["Vacant_Housing_Units"]
            sheet['C16'].value = self.dict_track["Median_Home_Value"]
            sheet['C18'].value = self.dict_track["Total_Hoouseholds"]
            sheet['C19'].value = self.dict_track["Avarage_Households_Size"]
            sheet['C20'].value = self.dict_track["Family_Households"]

            sheet['D7'].value = self.dict_zip_code["Total_Population"]
            sheet['D8'].value = self.dict_zip_code["Population_Growth_2010_2019"] + "(per year)"
            sheet['D9'].value = self.dict_zip_code["Population_Growth_2019_2024"] + "(per year)"
            sheet['D10'].value = self.dict_zip_code["Median_Household_Income"]
            sheet['D11'].value = self.dict_zip_code["Average_Household_Income"]
            sheet['D12'].value = self.dict_zip_code["Total_Housing_Units"]
            sheet['D13'].value = self.dict_zip_code["Owner_Occupied_HU"]
            sheet['D14'].value = self.dict_zip_code["Renter_Occupied_HU"]
            sheet['D15'].value = self.dict_zip_code["Vacant_Housing_Units"]
            sheet['D16'].value = self.dict_zip_code["Median_Home_Value"]
            sheet['D18'].value = self.dict_zip_code["Total_Hoouseholds"]
            sheet['D19'].value = self.dict_zip_code["Avarage_Households_Size"]
            sheet['D20'].value = self.dict_zip_code["Family_Households"]

            sheet['E7'].value = self.dict_city["Total_Population"]
            sheet['E8'].value = self.dict_city["Population_Growth_2010_2019"] + "(per year)"
            sheet['E9'].value = self.dict_city["Population_Growth_2019_2024"] + "(per year)"
            sheet['E10'].value = self.dict_city["Median_Household_Income"]
            sheet['E11'].value = self.dict_city["Average_Household_Income"]
            sheet['E12'].value = self.dict_city["Total_Housing_Units"]
            sheet['E13'].value = self.dict_city["Owner_Occupied_HU"]
            sheet['E14'].value = self.dict_city["Renter_Occupied_HU"]
            sheet['E15'].value = self.dict_city["Vacant_Housing_Units"]
            sheet['E16'].value = self.dict_city["Median_Home_Value"]
            sheet['E18'].value = self.dict_city["Total_Hoouseholds"]
            sheet['E19'].value = self.dict_city["Avarage_Households_Size"]
            sheet['E20'].value = self.dict_city["Family_Households"]

            sheet['F7'].value = self.dict_county["Total_Population"]
            sheet['F8'].value = self.dict_county["Population_Growth_2010_2019"] + "(per year)"
            sheet['F9'].value = self.dict_county["Population_Growth_2019_2024"] + "(per year)"
            sheet['F10'].value = self.dict_county["Median_Household_Income"]
            sheet['F11'].value = self.dict_county["Average_Household_Income"]
            sheet['F12'].value = self.dict_county["Total_Housing_Units"]
            sheet['F13'].value = self.dict_county["Owner_Occupied_HU"]
            sheet['F14'].value = self.dict_county["Renter_Occupied_HU"]
            sheet['F15'].value = self.dict_county["Vacant_Housing_Units"]
            sheet['F16'].value = self.dict_county["Median_Home_Value"]
            sheet['F18'].value = self.dict_county["Total_Hoouseholds"]
            sheet['F19'].value = self.dict_county["Avarage_Households_Size"]
            sheet['F20'].value = self.dict_county["Family_Households"]

            sheet['G7'].value = self.dict_metro["Total_Population"]
            sheet['G8'].value = self.dict_metro["Population_Growth_2010_2019"] + "(per year)"
            sheet['G9'].value = self.dict_metro["Population_Growth_2019_2024"] + "(per year)"
            sheet['G10'].value = self.dict_metro["Median_Household_Income"]
            sheet['G11'].value = self.dict_metro["Average_Household_Income"]
            sheet['G12'].value = self.dict_metro["Total_Housing_Units"]
            sheet['G13'].value = self.dict_metro["Owner_Occupied_HU"]
            sheet['G14'].value = self.dict_metro["Renter_Occupied_HU"]
            sheet['G15'].value = self.dict_metro["Vacant_Housing_Units"]
            sheet['G16'].value = self.dict_metro["Median_Home_Value"]
            sheet['G18'].value = self.dict_county["Total_Hoouseholds"]
            sheet['G19'].value = self.dict_county["Avarage_Households_Size"]
            sheet['G20'].value = self.dict_county["Family_Households"]

            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Dictionaries was completed & saved in {}".format(self.xls_name))
            logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
            return True
        except:
            print('Failed to copy to xls ')
            return False

    def return_dict_block(self):
        return self.dict_block

    def return_dict_basic_info(self):
        return self.dict_basic_info

    def return_dict_track(self):
        return self.dict_track

    def return_dict_zip_code(self):
        return self.dict_zip_code

    def return_dict_city(self):
        return self.dict_city

    def return_dict_county(self):
        return self.dict_county

    def return_dict_metro(self):
        return self.dict_metro

    def return_county_name(self):
        return self.county

    def return_zip_code_for_zillow_use(self):
        return self.dict_basic_info['zip_code']


class Crime(object):
    def __init__(self, street, state, city, short_state, xls_name):
        # all setup params
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.onboardnavigator_url = 'http://www.onboardnavigator.com/webcontent/OBWC_Search.aspx?&AID=102'
        self.city_data_url = 'http://www.city-data.com'
        self.home_facts_url = 'https://www.homefacts.com/'
        self.neighborhoodscout_url = 'https://www.neighborhoodscout.com/' + self.short_state.lower() + '/' + self.city.lower() + '/crime'
        self.bestplaces_url = 'https://www.bestplaces.net/crime/city/' + self.state.lower() + '/' + self.city.lower()
        # add NA
        #dictionaries
        self.dict_crime_total = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'zip_code': 'NA',
            'metropolitan': 'NA',
            'link_google_maps': 'NA'
        }
        self.dict_onboardnavigator = {
            'Total personal': 'NA',
            'Total property': 'NA',
            'Total overall': 'NA',
            'Year': '2019',

        }
        self.dict_city_data = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Year': '2019',

        }
        self.dict_home_facts = {
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Year': '2019',

        }
        self.dict_offenders = {
            'offender1': 'NA',
            'offender2': 'NA',
            'offender3': 'NA',

        }
        self.dict_neighborhoodscout = {
            'Diagram': 'NA',
            'List of safe areas': 'NA',
        }
        self.dict_bestplaces = {
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')
# the functions below written in a working flow
# getting all the information and copy into dicts
    def onboardnavigator_to_dict(self):
        try:
            print('onboardnavigator')
            driver = self.driver
            driver.get(self.onboardnavigator_url)
            time.sleep(10)
            print('Navigator tool opened')
            # select state
            state = driver.find_element_by_xpath('//*[@id="ddlGenLookupStateID"]').click()
            time.sleep(5)
            Select(driver.find_element_by_tag_name('select')).select_by_visible_text(self.state)
            time.sleep(5)
            print('state selected')
            driver.find_element_by_xpath('//*[@id="tbGenSearch"]').send_keys(self.city)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="radGenCity"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="cmdGenSave"]').click()
            time.sleep(10)
            print('navigator address located')
            link = driver.current_url
            self.dict_onboardnavigator['Total personal'] = link
            self.dict_onboardnavigator['Total property'] = link
            self.dict_onboardnavigator['Total overall'] = link
            print('onboardnavigator params was copied to dictionary , success {}'.format(self.dict_onboardnavigator))
        except:
            print('failed to locate navigator')

    def city_data_to_dict(self):
        try:
            print('citydata')
            driver = self.driver
            driver.get(self.city_data_url)
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').send_keys(self.city + ' ' + self.state)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="search_bar_box"]/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,4100)")
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="sex-offenders"]/p')))
            # select city data elemnets and copy to dictionary
            self.dict_city_data['total info'] = driver.find_element_by_xpath('//*[@id="sex-offenders"]/p').text
            self.dict_city_data['Pic of graph'] = driver.current_url
            self.dict_crime_total['total info'] = self.dict_city_data['total info']
            self.dict_crime_total['Pic of graph'] = self.dict_city_data['Pic of graph']
            print('city_data total info copied {}'.format(self.dict_city_data))
        except:
            print('failed to locate city data elements')
        try:
            driver = self.driver
            self.dict_city_data['Crime Index city'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[15]').text
            self.dict_city_data['US avarage'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[1]').text
            self.dict_crime_total['Crime Index city'] = self.dict_city_data['Crime Index city']
            self.dict_crime_total['US avarage'] = self.dict_city_data['US avarage']
            print('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            logging.debug('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            return True
        except:
            self.dict_city_data['Crime Index city'] = 'Crime table not exists in city_data for this state'
            print('Crime table not exists in city_data for this state')
            logging.debug('fail')
            return False

    def home_facts_to_dict(self):
        try:
            print('homefacts')
            driver = self.driver
            driver.get(self.home_facts_url)
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            addr = driver.find_element_by_xpath('//*[@id="fulladdress"]')
            addr.click()
            time.sleep(3)
            addr.send_keys(self.full_addr)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="main-search-form"]/div/div/div/div[1]/span/button').click()
            time.sleep(10)
            element = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div[2]/div[1]/div[3]/ul/li[1]/span[4]/a')
            driver.execute_script("window.scrollTo(0,600)")
            time.sleep(3)
            element.click()
            print(driver.current_url)
            print('view crime statistics report')
            time.sleep(10)
            try:
                print('trying to click')
                driver.find_element_by_partial_link_text('view crime statistics report').click()
            except:
                print('trying to click with second option')
                driver.execute_script("window.scrollTo(0,2700)")
                time.sleep(7)
                driver.find_element_by_partial_link_text('view crime statistics report').click()

            time.sleep(10)
            print(driver.current_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="crimeScore"]/div[1]/div[4]')))
            self.dict_home_facts['Overall Score'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[4]').get_attribute('class')
            self.dict_home_facts['Overall score big num'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[2]').text
            self.dict_home_facts['Score small procents'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[3]').text
            self.dict_home_facts['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_home_facts['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_home_facts['Score small procents'] = self.dict_home_facts['Score small procents']
            print(self.dict_home_facts['Overall Score'])
            print(self.dict_home_facts['Overall score big num'])
            print(self.dict_home_facts['Score small procents'])
            self.dict_crime_total['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_crime_total['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_crime_total['Score small procents'] = self.dict_home_facts['Score small procents']
            print('dict_home_facts params was copied to dictionary , success {}'.format(self.dict_home_facts))
            print('dict_offenders params was copied to dictionary , success {}'.format(self.dict_offenders))
        except:
            print('failed to locate and copy from home facts')

    def neighborhoodscout_to_dict(self):
        try:
            print('neighborhoodscout')
            print(self.neighborhoodscout_url)
            data = requests.get(self.neighborhoodscout_url)
            time.sleep(5)
            soup = BeautifulSoup(data.content, 'html.parser')
            list = soup.find_all('script', type='application/ld+json')
            list = str(list)
            index_list_start = list.find('itemListOrder')
            new_list = list[index_list_start:]
            index_list_end = new_list.find('</script>')
            orig_list = new_list[:index_list_end]
            index1 = orig_list.find('[')
            orig_list = orig_list[index1:]
            index2 = orig_list.find(']')
            # list of safety places taken from HTML converted to string
            orig_list = orig_list[index1:index2]
            self.dict_neighborhoodscout['List of safe areas'] = orig_list
            self.dict_neighborhoodscout['Diagram'] = self.neighborhoodscout_url
            print('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            logging.debug('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            return True
        except:
            logging.debug('fail to connect or copy from neighborhoodscout')
            print('fail to connect or copy from neighborhoodscout')
            return False

    def bestplaces_to_dict(self):
        try:
            print('bestplaces')
            driver = self.driver
            driver.get(self.bestplaces_url)
            time.sleep(5)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]')))
            self.dict_bestplaces['Violent crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]').text
            self.dict_bestplaces['Property crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[2]').text

            # Photos and Maps
            driver.find_element_by_xpath('//*[@id="form1"]/div[5]/div/div/p[3]/a[3]/u').click()
            time.sleep(4)
            self.dict_bestplaces['Photos and Maps of the city'] = driver.current_url
            print('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))
            logging.debug('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))

            self.dict_crime_total['Violent crime & US average'] = self.dict_bestplaces['Violent crime & US average']
            self.dict_crime_total['Property crime & US average'] = self.dict_bestplaces['Property crime & US average']
            self.dict_crime_total['Photos and Maps of the city'] = self.dict_bestplaces['Photos and Maps of the city']

            return True
        except:
            logging.debug('fail to connect or copy from bestplaces')
            print('fail to connect or copy from bestplaces')
            return False

# print all dictionaries
    def printall(self):
        print('All dictionaries\n')
        pp = pprint.PrettyPrinter(indent=4)
        print(self.dict_basic_info)
        pp.pprint(self.dict_basic_info)
        print(self.dict_onboardnavigator)
        pp.pprint(self.dict_onboardnavigator)
        print(self.dict_city_data)
        pp.pprint(self.dict_city_data)
        print(self.dict_home_facts)
        pp.pprint(self.dict_home_facts)
        print(self.dict_offenders)
        pp.pprint(self.dict_offenders)
        print(self.dict_neighborhoodscout)
        pp.pprint(self.dict_neighborhoodscout)
        print(self.dict_bestplaces)
        pp.pprint(self.dict_bestplaces)
        return True
# returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info

    def return_dict_onboardnavigator(self):
        return self.dict_onboardnavigator

    def return_dict_city_data(self):
        return self.dict_city_data

    def return_dict_home_facts(self):
        return self.dict_home_facts

    def return_dict_offenders(self):
        return self.dict_offenders

    def return_dict_neighborhoodscout(self):
        return self.dict_neighborhoodscout

    def return_dict_bestplaces(self):
        return self.dict_bestplaces

    def return_dict_crime_total(self):
        return self.dict_crime_total

# copy all dictionaries to xls file
    def xls_new_sheet_create(self):
                wb = openpyxl.load_workbook(self.xls_name)
                if wb.sheetnames.count(self.full_addr[:25]) == 0:
                    example_sheet = wb["example"]
                    wb.copy_worksheet(example_sheet)
                    # print(wb.sheetnames)
                    new_sheet = wb['example Copy']
                    new_sheet.title = self.full_addr[:25]
                    # print(wb.sheetnames)
                    wb.save(self.xls_name)
                    print("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                    logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                    wb.close()
                    return True
                else:
                    print("address is already exists in database!, recopy new run info ")
                    logging.debug("address is already exists in database!, recopy new run info ")
                    return False

    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.full_addr[:25]]

            sheet['A3'].value = self.dict_basic_info['street']
            sheet['C3'].value = self.dict_basic_info['city']
            sheet['D3'].value = self.dict_basic_info['state']

            sheet['B24'].value = self.dict_onboardnavigator['Total personal']
            sheet['B25'].value = self.dict_onboardnavigator['Total property']
            sheet['B26'].value = self.dict_onboardnavigator['Total overall']
            sheet['B27'].value = self.dict_onboardnavigator['Year']

            sheet['B29'].value = self.dict_city_data['Crime Index city']
            sheet['B30'].value = self.dict_city_data['US avarage']
            sheet['B31'].value = self.dict_city_data['Pic of graph']
            sheet['B32'].value = self.dict_city_data['total info']

            sheet['B34'].value = self.dict_home_facts['Overall Score']
            sheet['B35'].value = self.dict_home_facts['Overall score big num']
            sheet['B36'].value = self.dict_home_facts['Score small procents']
            sheet['B37'].value = self.dict_offenders['offender1']
            sheet['B38'].value = self.dict_offenders['offender2']
            sheet['B39'].value = self.dict_offenders['offender3']
            sheet['B40'].value = self.dict_home_facts['Year']

            sheet['B42'].value = self.dict_neighborhoodscout['Diagram']
            sheet['B43'].value = self.dict_neighborhoodscout['List of safe areas']

            sheet['B45'].value = self.dict_bestplaces['Violent crime & US average']
            sheet['B46'].value = self.dict_bestplaces['Property crime & US average']
            sheet['B47'].value = self.dict_bestplaces['Photos and Maps of the city']

            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Elements saved in {}".format(self.xls_name))
            logging.debug("Elements saved in {}".format(self.xls_name))
            return True
        except:
            return False


class Schools(object):
    def __init__(self, street, state, city, short_state, xls_name, county_name, zip_code):
        # all setup params
        self.zip_code = zip_code
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.county = county_name
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.greatschools_url = 'https://www.greatschools.org/'
        self.schooldigger_url = 'https://www.schooldigger.com/'
        self.homefacts_url = 'https://www.homefacts.com/'
        self.niche_url = 'https://www.niche.com/?ref=k12'

        self.dict_schools_general = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA',
            'school - HF elementary name': 'NA',
            'school - HF elementary link': 'NA',
            'school - HF middle name': 'NA',
            'school - HF middle link': 'NA',
            'school - HF high name': 'NA',
            'school - HF high link': 'NA'

        }


        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'county': self.county
        }
        self.dict_greatschools = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_schooldigger = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'
        }
        self.dict_homefacts = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_niche = {
            'link - County Schools': 'NA',
            'name - global': 'NA',
            'rank - School Districts if exists': 'NA',
            'grade - overall niche grade': 'NA',
            'link - all ranks state county schools/metropolitan/national': 'NA'

        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def greateschools_to_dict(self):
            try:  # connecting to greateschools
                driver = self.driver
                driver.get(self.greatschools_url)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input')))
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').click()
                time.sleep(2)
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
                time.sleep(2)
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
                time.sleep(5)
                print(driver.current_url)
                try:
                    # elementary school assigned tags
                    time.sleep(10)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Elementary School')))
                    driver.find_element_by_partial_link_text('Elementary School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - elementary link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - elementary link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    self.dict_greatschools['school - elementary name'] = school_name
                    print('Elementary school name: {}'.format(school_name))
                    self.dict_schools_general['school - elementary link'] = self.dict_greatschools['school - elementary link']
                    self.dict_schools_general['school - elementary name'] = self.dict_greatschools['school - elementary name']
                    driver.back()
                except:
                    print('failed to locate elemantary school from greateschools')

                try:
                    # middle school assigned tags
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Middle School')))
                    driver.find_element_by_partial_link_text('Middle School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - middle link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - middle link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    print('Middle school name: {}'.format(school_name))
                    self.dict_greatschools['school - middle name'] = school_name
                    self.dict_schools_general['school - middle link'] = self.dict_greatschools['school - middle link']
                    self.dict_schools_general['school - middle name'] = self.dict_greatschools['school - middle name']
                    driver.back()
                except:
                    print('failed to locate middle school from greateschools')
                    self.dict_greatschools['school - middle link'] = 'NA'


                try:
                    # high school assigned tags
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'High School')))
                    driver.find_element_by_partial_link_text('High School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - high link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - high link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    print('High school name: {}'.format(school_name))
                    self.dict_greatschools['school - high name'] = school_name
                    self.dict_schools_general['school - high link'] = self.dict_greatschools['school - high link']
                    self.dict_schools_general['school - high name'] = self.dict_greatschools['school - high name']
                    driver.back()
                except:
                    print('failed to locate high school from greateschools')
            except:
                print('something went wrong with greateschools')

    def schooldigger_to_dict(self): #check
        try:
            driver = self.driver
            driver.get(self.schooldigger_url)
            print(driver.current_url)
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtHPAC"]')))
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(Keys.ENTER)
            time.sleep(3)
            print(driver.current_url)

            # elementary schools under boundary tags
            try:
                driver.find_element_by_partial_link_text('Elementary').click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - elementary link'] = driver.current_url
                self.dict_schooldigger['school - elementary name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                print('elemantary school found in schooldigger')
                driver.back()
                time.sleep(10)
            except:
                print('elemantary school was not fount in schooldigger')

            try:
                driver.find_element_by_partial_link_text('Middle').click()
                time.sleep(5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - middle link'] = driver.current_url
                self.dict_schooldigger['school - middle name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                driver.back()
                time.sleep(10)
                print('middle school found in schooldigger')
            except:
                print('middle school was not fount in schooldigger')
            # high boundary tags
            try:
                driver.find_element_by_partial_link_text('High').click()
                time.sleep(10)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - high link'] = driver.current_url
                self.dict_schooldigger['school - high name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                print('high school found in schooldigger')
                driver.back()
                time.sleep(10)
            except:
                print('high school was not fount in schooldigger')
            print('schooldigger params was copied to dictionary , success {}'.format(self.dict_schooldigger))
        except:
            print('failed to connect or locate params from schooldigger')

    def homefacts_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.homefacts_url)
            time.sleep(10)
            print('homefacts entered ')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            driver.find_element_by_xpath('//*[@id="fulladdress"]').click()
            time.sleep(2)
            # driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(str(self.zip_code))
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(Keys.ENTER)
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="navbar"]/ul/li[4]/a')))
            driver.find_element_by_xpath('//*[@id="navbar"]/ul/li[4]/a').click()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,550)")
            time.sleep(3)
            print(driver.current_url)
            time.sleep(3)
            print('schools list located')
            # elementary
            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'ELEMENTARY SCHOOL')))
                driver.find_element_by_partial_link_text('ELEMENTARY SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('ELEMENTARY {}'.format(school_name))
                time.sleep(3)
                self.dict_homefacts['school - elementary name'] = school_name #                 //*[@id="school_year_2019"]/div[1]/div[2]
                self.dict_homefacts['school - elementary link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - elementary link'])
                time.sleep(2)
                self.dict_schools_general['school - HF elementary name'] = self.dict_homefacts['school - elementary name']
                self.dict_schools_general['school - HF elementary link'] = self.dict_homefacts['school - elementary link']
                print('elemantary school found in homefacts')
                driver.back()
                time.sleep(5)
            except:
                print('elemantary school was not fount in homefacts')

            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'MIDDLE SCHOOL')))
                driver.find_element_by_partial_link_text('MIDDLE SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('middle {}'.format(school_name))
                self.dict_homefacts['school - middle name'] = school_name
                self.dict_homefacts['school - middle link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - middle link'])
                time.sleep(2)
                self.dict_schools_general['school - HF middle name'] = self.dict_homefacts['school - middle name']
                self.dict_schools_general['school - HF middle link'] = self.dict_homefacts['school - middle link']
                driver.back()
                time.sleep(5)
            except:
                print('middle school was not fount in homefacts')

            try:
                print('trying to locate high school')
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'HIGH SCHOOL')))
                driver.find_element_by_partial_link_text('HIGH').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print('High school {}'.format(school_name))
                time.sleep(5)
                self.dict_homefacts['school - high name'] = school_name
                print('trying to locate high school grade from pic') # //*[@id="school_year_2018"]/div[1]/div[2]
                self.dict_homefacts['school - high link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - high link'])
                time.sleep(2)
                self.dict_schools_general['school - HF high name'] = self.dict_homefacts['school - high name']
                self.dict_schools_general['school - HF high link'] = self.dict_homefacts['school - high link']
                driver.back()
                time.sleep(5)
            except:
                print('high school was not fount in homefacts')

            print('homefacts params was copied to dictionary , success {}'.format(self.dict_homefacts))
        except:
            print('fail to copy params from homefacts')

    def niche_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.niche_url)
            time.sleep(4)
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').click()
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').send_keys(
                self.county + ' county')
            time.sleep(1)
            # driver.find_element_by_xpath('//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').sendKeys(Keys.ENTER)
            time.sleep(4)
            self.dict_niche['link - County Schools'] = driver.current_url
            self.dict_niche['name - global'] = driver.current_url
            self.dict_niche['rank - School Districts if exists'] = driver.current_url
            self.dict_niche['grade - overall niche grade'] = driver.current_url
            self.dict_niche['link - all ranks state county schools/metropolitan/national'] = driver.current_url

            print('niche params was copied to dictionary , success {} '.format(self.dict_niche))
            return True
        except:
            print('fail to locate params from niche')
            logging.debug('fail')
            return False

    def printall(self):
        pp = pprint.PrettyPrinter(indent=4)
        print('greate schools')
        pp.pprint(self.dict_greatschools)
        print('school digger')
        pp.pprint(self.dict_schooldigger)
        print('home facts')
        pp.pprint(self.dict_homefacts)
        print('niche')
        pp.pprint(self.dict_niche)

# returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info
    def return_dict_greateshcools(self):
        return self.dict_greatschools
    def return_dict_schooldigger(self):
        return self.dict_schooldigger
    def return_dict_homefacts(self):
        return self.dict_homefacts
    def return_dict_niche(self):
        return self.dict_niche
    def return_dict_schools_general(self):
        return self.dict_schools_general
# copy all dictionaries to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.full_addr[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.full_addr[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet name: {}".format(self.full_addr[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database recopy new params")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        print('copy dicts to xls')
        wb = openpyxl.load_workbook(self.xls_name)
        sheet = wb[self.full_addr[:25]]
        # print(wb.sheetnames)
        sheet['F24'].value = self.dict_greatschools['school - elementary name']
        sheet['F25'].value = self.dict_greatschools['school - elementary link']
        sheet['F26'].value = self.dict_greatschools['school - middle name']
        sheet['F27'].value = self.dict_greatschools['school - middle link']
        sheet['F28'].value = self.dict_greatschools['school - high name']
        sheet['F29'].value = self.dict_greatschools['school - high link']

        sheet['F31'].value = self.dict_schooldigger['school - elementary name']
        sheet['F32'].value = self.dict_schooldigger['school - elementary link']
        sheet['F33'].value = self.dict_schooldigger['school - middle name']
        sheet['F34'].value = self.dict_schooldigger['school - middle link']
        sheet['F35'].value = self.dict_schooldigger['school - high name']
        sheet['F36'].value = self.dict_schooldigger['school - high link']

        sheet['F38'].value = self.dict_homefacts['school - elementary name']
        sheet['F39'].value = self.dict_homefacts['school - elementary link']
        sheet['F40'].value = self.dict_homefacts['school - middle name']
        sheet['F41'].value = self.dict_homefacts['school - middle link']
        sheet['F42'].value = self.dict_homefacts['school - high name']
        sheet['F43'].value = self.dict_homefacts['school - high link']

        # sheet['F29'].value = self.dict_basic_info['street']
        # sheet['F30'].value = self.dict_basic_info['city']
        # sheet['F31'].value = self.dict_basic_info['state']

        wb.save(self.xls_name)
        wb.close()
        # printing the process
        print("Dictionaries was completed & saved in {}".format(self.xls_name))
        logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
        return True


class Builders(object):
    def __init__(self, metropolitan, short_state, xls_name):
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")
        self.lennar_url = 'https://www.lennar.com/'  # builders website
        self.metropolitan = metropolitan.lower() + ' ' + short_state.lower()  # full name for search
        self.floorplan_homes = ''
        self.xls_name = xls_name  # xls name
        self.short_state = short_state
        self.clicked = ''
        self.list_of_homes = []
        self.community_address_list_full = []  # full list of community addresses
        self.community_address_list_names = []  # full list of community names
        self.id_random_list = []
        self.row = 2
        self.rowhome = 2
        self.general_row = 0
        self.row_num_xls = 0
        self.index = 1
        self.x_path_name_to_scroll = ''
        self.element = ''  # scrolling element
        self.addr = ''
        self.name = ''
        self.update_time = ''  # update time
        self.homes_urls = []  # list of homes urls
        self.x_path_name = ''
        self.num_of_communities = ''
        self.num_of_pages = ''  # num of community pages
        self.num_of_comm_pages = ''
        self.num_of_homes_pages = ''  # num of homes pages
        self.num_of_moving_homes = ''  # num of homes
        self.dict_lennar_filter_info = {
            'Communities num': '',
            'metropolitan name': metropolitan,
            'Quick Move-In Homes num': '',
            'Floorplans num': '',
            'time of update': ''
        }

        # community data for mysql and xls
        self.dict_community_data = {
            'address': '',
            'name_community': '',
            'overview': '',
            'approximate_hoa_fees': 'non',
            'approximate_tax_rate': 'non',
            'included_features_pdf_url': 'under solution',
            'community_map_url': 'no pic',
            'community_home_picture_for_present_url': 'no pic',
            'available_homes_quick_move_in_homes': '',
            'available_homes_floorplans': '',
            'id_generated':''

        }

        # home data for mysql and xls
        self.dict_home_data = {
            'address': '',
            'name_community': metropolitan,
            'home_name': '',
            'home_site': '',
            'availability': '',
            'priced_from': '',
            'home_size': '',
            'stories': '',
            'beds': '',
            'type': '',
            'baths': '',
            'garage': '',
            'id': '',
            'id_generated': '',
            'description': '',
            'included_features_pdf': 'under solution',
            'floorplans_with_furniture_pic': '',
            'id_generated_home': '',
            'gallery_view_picture': ''
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def lennar_filter_and_toolbar_info_copy(self):
        driver = self.driver
        driver.get(self.lennar_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input')))  # await command
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').click()
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').send_keys(self.metropolitan)
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/button').send_keys(Keys.ENTER)
        print('Connected to Lennar')
        time.sleep(3)
        # Create filter
        print('creating filter')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]')))
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]').click()
        time.sleep(3)
        # community type
        try:
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/a').click()
            time.sleep(3)
        except:
            print('no community type')
        # add single family loop
        for i in range(0, 10):
            print(i)
            element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div/div/ul/li[' + str(i) + ']/label'
            try:
                filter = driver.find_element_by_xpath(element).text
            except:
                print('not such element exists')
                filter = 'NO'

            if filter == 'Single Family':
                mainelem = element
                filter = driver.find_element_by_xpath(mainelem)
                filter.click()
                print(filter.text)
                time.sleep(3)
                print('Applied Filters ,success')
                break
            else:
                print('element not found on {}'.format(i))

        # select price
        print('selecting price')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a')))
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a').click()
        time.sleep(3)

        try:
            # set price < 300$
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/span').click()
            time.sleep(3)
            for i in range(0, 10):
                print(i)
                element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/ul/li[' + str(i) + ']'
                print(element)
                try:
                    filter = driver.find_element_by_xpath(element).text
                except:
                    print('not such element exists')
                    filter = 'NO'

                if filter == '300K':  #  300K
                    mainelem = element
                    filter = driver.find_element_by_xpath(mainelem)
                    filter.click()
                    print(filter.text)
                    time.sleep(3)
                    print('Applied Filters ,success')
                    time.sleep(3)
                    print('clicking on botton')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[3]/div/div/a[2]').click()
                    break
                else:
                    print('element not found on {}'.format(i))
        except:
            print('price element not found')

        print('locating basic info about communities')
        time.sleep(3)
        try:
            driver = self.driver
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a')))
            self.dict_lennar_filter_info['Communities num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a').text
            index1 = self.dict_lennar_filter_info['Communities num'].find('(')
            index2 = self.dict_lennar_filter_info['Communities num'].find(')')
            self.dict_lennar_filter_info['Communities num'] = self.dict_lennar_filter_info['Communities num'][index1 + 1:index2]
            self.dict_lennar_filter_info['Quick Move-In Homes num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[2]/a').text
            self.dict_lennar_filter_info['Floorplans num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[3]/a').text
            self.num_of_communities = self.dict_lennar_filter_info['Communities num']
            print('Communities number is {}'.format(self.dict_lennar_filter_info['Communities num']))
            print('change view to list')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a')))
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
            time.sleep(3)
            print('Basic Information scanned {}'.format(self.dict_lennar_filter_info))
        except:
            print('Failed, to Locate information from Lennar')

        try:
            print('copy to xls basic info')
            print('Creating new xls sheet')
            wb = openpyxl.load_workbook(self.xls_name)
            if wb.sheetnames.count(self.metropolitan + ' comm_data') == 0:
                example_sheet = wb['comm_data']
                wb.copy_worksheet(example_sheet)
                new_sheet = wb['comm_data Copy']
                new_sheet.title = self.metropolitan + ' comm_data'
                wb.save(self.xls_name)
                print("XLS new sheet is ready, sheet name: {}".format(new_sheet.title))
                wb.close()
            else:
                print('address was exist in xls')
        except:
            print('failed to connect to xls')

        try:
            time.sleep(3)
            print('opening xls')
            print('xls name {}'.format(self.xls_name))
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.metropolitan + ' comm_data']
            sheet['K2'].value = self.dict_lennar_filter_info['metropolitan name']
            sheet['L2'].value = self.dict_lennar_filter_info['Communities num']
            sheet['M2'].value = self.dict_lennar_filter_info['Quick Move-In Homes num']
            sheet['N2'].value = self.dict_lennar_filter_info['Floorplans num']
            sheet['J2'].value = datetime.datetime.now()
            wb.save(self.xls_name)
            wb.close()
            print('sheet name is {}'.format(self.metropolitan + ' comm_data'))
            print('basic community info bar was saved in xls')
            return True
        except:
            print('failed to copy basic community info to XLS ')
            logging.debug('failed to open XLS')
            return False

    '''
        def community_and_homes_all_data_to_xls_and_SQL(self):
        try:
            print('Calculating the num of Pages to scroll - communities')
            if int(self.num_of_communities) < 30:
                self.num_of_comm_pages = 1
                print('Num of communities {}'.format(self.num_of_communities))
                print('Num of pages of communities {}'.format(self.num_of_comm_pages))
            else:
                if int(self.num_of_communities) < 60:
                    self.num_of_comm_pages = 2
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
                else:
                    self.num_of_comm_pages = int(self.num_of_communities) / 30
                    self.num_of_comm_pages = round(self.num_of_comm_pages)
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
        except:
            print('could not calculate data about communities')

        if int(self.num_of_communities) < 30:  # if communities < 30 (one page)
            print('if communities < 30')
            time.sleep(2)
            for x in range(0, int(self.num_of_communities)):  # int(self.num_of_communities):
                print('community area entered')
                try:
                    driver = self.driver
                    time.sleep(5)
                    print('change view to list')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                    time.sleep(5)
                except:
                    print('list button not located')
                try:
                    driver = self.driver
                    print('Preparing to Enter community on num {}'.format(x))
                    time.sleep(10)
                    print('trying to locate community address')
                    x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[2]'
                    print(x_path)
                    self.addr = driver.find_element_by_xpath(x_path).text
                    self.dict_community_data['address'] = self.addr
                    print('Community Address: {}'.format(self.addr))
                    self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[1]/a/strong'
                    self.name = driver.find_element_by_xpath(self.x_path_name).text
                    self.dict_community_data['name_community'] = self.name
                    print('Community Name: {}'.format(self.name))
                    self.community_address_list_full.append(self.dict_community_data['address'])
                    print('Community address was added to list for automation')
                    print('scrolling')
                    scroll = 245 * x
                    print(scroll)
                    scroll = "window.scrollTo(0, " + str(scroll) + ")"
                    driver.execute_script(scroll)
                    time.sleep(10)
                    print('scrolled'.format(x))
                    print('trying to click the scrolled community')
                    print(driver.current_url)
                    print(self.x_path_name)
                    print('clicking')
                    driver.find_element_by_xpath(self.x_path_name).click()
                    time.sleep(3)
                    print('clicked')
                    time.sleep(10)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                    print('SUCCESS - community found and pressed')
                    time.sleep(10)
                    c = 0
                except:
                    print('FAILED - to locate community on xpath num {}'.format(x))
                    c = 1

                # if community is located
                if c == 0:
                    print('After Community was located - starting to download data')
                    print('1 First generating ID for Community')
                    time.sleep(5)
                    self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                    self.id_random_list.append(self.dict_community_data['id_generated'])
                    print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                    try:
                        print('2 try copy overview data')
                        driver = self.driver
                        self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                    except:
                        print('failed to locate overview')

                    time.sleep(5)
                    try:
                        print('3 try copy picture 1 map ')
                        driver = self.driver
                        self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                        print('actually downloading the image and changing the name.jpg')
                        urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                    except:
                        print('failed to locate pictures map')
                        self.dict_community_data['community_map_url'] = 'NA'
                    try:
                        print('4 try copy pictures 2')
                        driver = self.driver
                        self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                        urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                    except:
                        print('failed to locate pictures 2')
                        self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                    try:
                        print('5 Available Homes and floorplans')
                        driver = self.driver
                        self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                        self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                        print('success to copy home toolbar data {}'.format(self.dict_community_data))
                    except:
                        print('failed to locate homes toolbar')

                    try:
                        print('6 copy community data to xls')
                        print('open xls '.format(self.xls_name))
                        wb = openpyxl.load_workbook(self.xls_name)
                        time.sleep(2)
                        sheet = wb[self.metropolitan + ' comm_data']
                        sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                        sheet['B' + str(self.row)].value = self.dict_community_data['address']
                        sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                        sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                        sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                        sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                        sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                        sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                        sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                        wb.save(self.xls_name)
                        wb.close()
                        print('COMMUNITY DATA - saved in xls')
                        self.row = self.row + 1
                    except:
                        print('failed to copy community data to XLS ')

                    # homes general data
                    try:
                        driver = self.driver
                        print('scrolling to homes')
                        driver.execute_script("window.scrollTo(0, 2050)")
                        time.sleep(5)
                        print('changing view to list')
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                        print('Calculating num of homes')
                        self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                        print('num of homes to verify {}'.format(self.num_of_moving_homes))
                        print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans'][12:-1]))
                        self.floorplan_homes = self.dict_community_data['available_homes_floorplans'][12:-1]
                        time.sleep(3)
                    except:
                        print('could not locate general homes information')

                    print('copy homes + floorpans :):):):):):)')
                    print('FLOORPLANS')
                    for j in range(0, int(self.floorplan_homes)):
                        try:
                            driver = self.driver
                            print('Choosing floorplans Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                            print('floorplans clicked')
                            time.sleep(5)
                            print('For floorplan - Home number {}'.format(j + 1))
                            time.sleep(3)
                            print('Scrolling to Home')
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to floorplans Home')
                            time.sleep(3)
                            self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                        except:
                            print('could not locate floorplan home!')

                        print('floorplans - trying to enter - Homes')

                        if int(self.floorplan_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if floorplans home is <= 1 , trying to find home link')
                                print('clicking on floorplans home link')
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans home entered')
                                print('waiting for the floorplans home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans Home LOCATED in the list!')
                                time.sleep(15)
                                print('first generating ID floorplans for home')
                                time.sleep(5)
                                print('generating home floorplans id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans could not locate home link <= 1')
                        else:
                            try:
                                driver = self.driver
                                print('IF floorplans Homes count more than > 1')
                                print('clicking on floorplans home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home floorplan LOCATED!')
                                time.sleep(15)
                                print('generating floorplans home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans Home not located on path number {}'.format(j + 1))

                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "TBB"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                            time.sleep(2)
                        except:
                            print('failed to generate')

                        try:
                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                self.dict_home_data['availability'] = 'NA'
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')

                            # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                        except:
                            print('could not locate HOME / elements')

                        print('Trying to copy all gained Homes data to XLS file')
                        try:
                            print('xls - creating new sheet with home name')
                            wb = openpyxl.load_workbook(self.xls_name)
                            if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                                print('creating xls')
                                example_sheet = wb['home_data']
                                wb.copy_worksheet(example_sheet)
                                print(wb.sheetnames)
                                new_sheet = wb['home_data Copy']
                                new_sheet.title = self.metropolitan + ' home_data'
                                wb.save(self.xls_name)
                                print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                                print(wb.sheetnames)
                                wb.close()
                            else:
                                print("Metropolitan Homes sheet already created in xls")
                        except:
                            print('failed to connect to xls file and create sheet')

                        # copy home basic info to xls
                        try:
                            # opening xls
                            print('IMPORTANT - copy home info to xls')
                            wb = openpyxl.load_workbook(self.xls_name)
                            sheet = wb[self.metropolitan + ' home_data']
                            sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                            sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                            sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                            sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                            sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                            sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                            sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                            sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                            sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                            sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                            sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                            sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                            sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                            sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                            sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                            sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                            sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                            sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                            sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']

                            wb.save(self.xls_name)
                            wb.close()
                            print('xls floorplan - HOME params was saved')
                            self.rowhome = self.rowhome + 1
                        except:
                            print('failed to copy floorplans HOME params to xls')
                            logging.debug('failed to open XLS')

                        print('Trying to Connect and copy same data to MySQL server')
                        self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                        try:
                            db = mysql.connector.connect(
                                host='107.180.21.18',
                                user='grow097365',
                                passwd='Jknm678##Tg',
                                database='equity_property'
                            )
                            mycursor = db.cursor()
                            print(db)  # checking our connection to DB
                            command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                            print(command)

                            mycursor.execute(command)
                            myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                            print(len(myresult))
                            print(myresult)

                            if len(myresult) == 0:
                                print('Similar homes not found, copying to database!')
                                db = mysql.connector.connect(
                                    host='107.180.21.18',
                                    user='grow097365',
                                    passwd='Jknm678##Tg',
                                    database='equity_property'
                                )
                                mycursor = db.cursor()
                                print(db)
                                sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                val = (self.dict_home_data['id_generated'],
                                       datetime.datetime.now(),
                                       self.dict_home_data['address'],
                                       self.short_state,
                                       self.metropolitan,
                                       self.dict_home_data['home_name'],
                                       self.dict_home_data['home_size'],
                                       self.dict_home_data['beds'],
                                       self.dict_home_data['baths'],
                                       self.dict_home_data['garage'],
                                       self.dict_home_data['priced_from'],
                                       self.dict_home_data['gallery_view_picture'],
                                       self.dict_home_data['type'],
                                       str(self.dict_home_data['id_generated_home']),
                                       self.dict_home_data['name_community'])
                                mycursor.execute(sql, val)
                                db.commit()
                                time.sleep(3)
                                print('IMPORTANT - Home floorplan data copied to mySQL')
                            else:
                                print('Similar home found in database')
                        except:
                            print('failed to work with mySQL')

                        try:
                            driver = self.driver
                            print('trying to go back to HOMES list after data copied')
                            driver.back()
                            time.sleep(7)
                        except:
                            print('could not go back on general HOMES list')
                    print('HOMES')
                    for j in range(0, int(self.num_of_moving_homes)):
                        try:
                            driver = self.driver
                            print('entering Homes and copy the data')
                            print('For Home number {}'.format(j + 1))
                            print('Choosing quick mov in Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                            time.sleep(3)
                            print('Scrolling to Home')
                            time.sleep(3)
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to Homes')
                            print('trying to enter - Homes')
                        except:
                            print('could not locate floorplan home!')

                        if int(self.num_of_moving_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if home is <= 1 , trying to find home link')
                                print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('home link clicked')
                                time.sleep(5)
                                print('home entered')
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('could not locate home')
                        else:
                            try:
                                driver = self.driver
                                print('if Homes more then > 1')
                                print('clicking on home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                time.sleep(5)
                                print('home link clicked')
                                print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('Home not located on path number {}'.format(j + 1))
                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "MIR"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))

                            time.sleep(2)
                            try:
                                print('try copy home picture')
                                driver = self.driver
                                print('getting the source link of the picture')
                                self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                                print(self.dict_home_data['gallery_view_picture'])
                                urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                            except:
                                print('failed to locate pictures')
                                self.dict_home_data['gallery_view_picture'] = 'NA'

                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                driver = self.driver
                                self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                                print(self.dict_home_data['baths'])
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')
                        except:
                            print('could not locate HOME / elements')
                    print('after all homes was scanned, we going back to community')
                    try:
                        driver = self.driver
                        time.sleep(5)
                        driver.back()
                        time.sleep(5)
                        driver.back()
                        time.sleep(10)
                        print('Waiting till the page will load the community')
                    except:
                        print('could not go back on community list')
            print('END of work on communities < 30')
        else:
            print('if communities > 30 and we got pages to scroll')
            time.sleep(2)
            for page in range(self.num_of_comm_pages):
                print('Comm page num {}'.format(page + 1))
                for self.row_num_xls in range(0, 29):  # 30 communities per page
                    print('community area entered')
                    try:
                        driver = self.driver
                        print('change view to list')
                        time.sleep(5)
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                    except:
                        print('list button not located')
                    try:
                        driver = self.driver
                        print('Preparing to Enter community on num {}'.format(self.row_num_xls))
                        time.sleep(10)
                        print('trying to locate community address')
                        x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[2]'
                        print(x_path)
                        self.addr = driver.find_element_by_xpath(x_path).text
                        self.dict_community_data['address'] = self.addr
                        print('Community Address: {}'.format(self.addr))
                        self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[1]/a/strong'
                        self.name = driver.find_element_by_xpath(self.x_path_name).text
                        self.dict_community_data['name_community'] = self.name
                        print('Community Name: {}'.format(self.name))
                        self.community_address_list_full.append(self.dict_community_data['address'])
                        print('Community address was added to list for automation')

                        print('scrolling')
                        scroll = 245 * self.row_num_xls
                        print(scroll)
                        scroll = "window.scrollTo(0, " + str(scroll) + ")"
                        driver.execute_script(scroll)
                        time.sleep(10)
                        print('scrolled'.format(self.row_num_xls))

                        print('trying to click the scrolled community')
                        print(driver.current_url)
                        print(self.x_path_name)
                        print('clicking')
                        driver.find_element_by_xpath(self.x_path_name).click()
                        time.sleep(3)
                        print('clicked')
                        time.sleep(10)
                        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        print('SUCCESS - community found and pressed')
                        time.sleep(10)
                        c = 0
                    except:
                        print('FAILED - to locate community on xpath num {}'.format(self.row_num_xls))
                        c = 1

                    # if community is located
                    if c == 0:
                        print('After Community was located - starting to download data')
                        print('1 First generating ID for Community')
                        time.sleep(5)
                        self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                        self.id_random_list.append(self.dict_community_data['id_generated'])
                        print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                        try:
                            print('2 try copy overview data')
                            driver = self.driver
                            self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                        except:
                            print('failed to locate overview')

                        time.sleep(5)
                        try:
                            print('3 try copy picture 1 map ')
                            driver = self.driver
                            self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                            print('actually downloading the image and changing the name.jpg')
                            urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                        except:
                            print('failed to locate pictures map')
                            self.dict_community_data['community_map_url'] = 'NA'
                        try:
                            print('4 try copy pictures 2')
                            driver = self.driver
                            self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                        except:
                            print('failed to locate pictures 2')
                            self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                        try:
                            print('5 Available Homes and floorplans')
                            driver = self.driver
                            self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                            self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                                                                                                                 # //*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]
                            print('success to copy home toolbar data {}'.format(self.dict_community_data))
                        except:
                            print('failed to locate homes toolbar')

                        try:
                            print('6 copy community data to xls')
                            print('open xls '.format(self.xls_name))
                            wb = openpyxl.load_workbook(self.xls_name)
                            time.sleep(2)
                            sheet = wb[self.metropolitan + ' comm_data']
                            sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                            sheet['B' + str(self.row)].value = self.dict_community_data['address']
                            sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                            sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                            sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                            sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                            sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                            sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                            sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                            wb.save(self.xls_name)
                            wb.close()
                            print('COMMUNITY DATA - saved in xls')
                            self.row = self.row + 1
                        except:
                            print('failed to copy community data to XLS ')

                        try:
                            print('# homes general data')
                            driver = self.driver
                            print('scrolling to homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(5)
                            print('changing view to list')
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                            time.sleep(5)
                            print('Calculating num of homes')
                            self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                            self.floorplan_homes = self.dict_community_data['available_homes_floorplans'][-2:-1]
                            print('num of homes to verify {}'.format(self.num_of_moving_homes))
                            print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans']))
                            time.sleep(3)
                        except:
                            print('could not locate general homes information')

                        print('copy homes + floorpans :):):):):):)')
                        print('FLOORPLANS')
                        for j in range(0, int(self.floorplan_homes)):
                            try:
                                driver = self.driver
                                print('Choosing floorplans Homes')
                                driver.execute_script("window.scrollTo(0, 2050)")
                                time.sleep(3)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                                print('floorplans clicked')
                                time.sleep(5)
                                print('For floorplan - Home number {}'.format(j + 1))
                                time.sleep(3)
                                print('Scrolling to Home')
                                scroll = 2000 + (245 * j)
                                scroll = "window.scrollTo(0, " + str(scroll) + ")"
                                driver.execute_script(scroll)
                                print('scrolled to floorplans Home')
                                time.sleep(3)
                                self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                                urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                            except:
                                print('could not locate floorplan home!')

                            print('floorplans - trying to enter - Homes')

                            if int(self.floorplan_homes) <= 1:
                                try:
                                    driver = self.driver
                                    print('if floorplans home is <= 1 , trying to find home link')
                                    print('clicking on floorplans home link')
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('floorplans home link clicked')
                                    time.sleep(5)
                                    print('floorplans home entered')
                                    print('waiting for the floorplans home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('floorplans Home LOCATED in the list!')
                                    time.sleep(15)
                                    print('first generating ID floorplans for home')
                                    time.sleep(5)
                                    print('generating home floorplans id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('floorplans could not locate home link <= 1')
                            else:
                                try:
                                    driver = self.driver
                                    print('IF floorplans Homes count more than > 1')
                                    print('clicking on floorplans home link')
                                    ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                    time.sleep(5)
                                    print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                    print('floorplans home link clicked')
                                    time.sleep(5)
                                    print('floorplans waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home floorplan LOCATED!')
                                    time.sleep(15)
                                    print('generating floorplans home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('floorplans Home not located on path number {}'.format(j + 1))

                            try:
                                self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                                self.dict_home_data['type'] = "TBB"
                                print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                                time.sleep(2)
                            except:
                                print('failed to generate')

                            try:
                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                    print(self.dict_home_data['home_name'])
                                except:
                                    print('home name not found')

                                try:
                                    self.dict_home_data['address'] = self.dict_community_data['address']
                                    print(self.dict_home_data['address'])
                                except:
                                    print('address not found')

                                try:
                                    self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                    print(self.dict_home_data['name_community'])
                                except:
                                    print('name community not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['home_site'])
                                except:
                                    print('home site not found')

                                self.dict_home_data['included_features_pdf'] = 'under solution'

                                try:
                                    self.dict_home_data['availability'] = 'NA'
                                    print(self.dict_home_data['availability'])
                                except:
                                    print('availability not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                    self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                    print(self.dict_home_data['priced_from'])
                                except:
                                    print('priced from not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['home_size'])
                                except:
                                    print('home size not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                    print(self.dict_home_data['stories'])
                                except:
                                    print('stories not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                    print(self.dict_home_data['beds'])
                                except:
                                    print('beds not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                    print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                                except:
                                    print('baths not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                    print(self.dict_home_data['garage'])
                                except:
                                    print('garage not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                    print(self.dict_home_data['description'])
                                except:
                                    print('description not found')

                                try:
                                    driver = self.driver
                                    print('trying to copy home FloorPlan Pic scrolling')
                                    driver.execute_script("window.scrollTo(0, 1600)")
                                    time.sleep(4)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                    time.sleep(3)
                                    self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                                except:
                                    print('could not locate home pics and FloorPlan Pic')

                                # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                            except:
                                print('could not locate HOME / elements')

                            print('Trying to copy all gained Homes data to XLS file')
                            try:
                                print('xls - creating new sheet with home name')
                                wb = openpyxl.load_workbook(self.xls_name)
                                if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                                    print('creating xls')
                                    example_sheet = wb['home_data']
                                    wb.copy_worksheet(example_sheet)
                                    print(wb.sheetnames)
                                    new_sheet = wb['home_data Copy']
                                    new_sheet.title = self.metropolitan + ' home_data'
                                    wb.save(self.xls_name)
                                    print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                                    print(wb.sheetnames)
                                    wb.close()
                                else:
                                    print("Metropolitan Homes sheet already created in xls")
                            except:
                                print('failed to connect to xls file and create sheet')

                            # copy home basic info to xls
                            try:
                                # opening xls
                                print('IMPORTANT - copy home info to xls')
                                wb = openpyxl.load_workbook(self.xls_name)
                                sheet = wb[self.metropolitan + ' home_data']
                                sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                                sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                                sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                                sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                                sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                                sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                                sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                                sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                                sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                                sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                                sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                                sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                                sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                                sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                                sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                                sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                                sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                                sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                                sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']

                                wb.save(self.xls_name)
                                wb.close()
                                print('xls floorplan - HOME params was saved')
                                self.rowhome = self.rowhome + 1
                            except:
                                print('failed to copy floorplans HOME params to xls')
                                logging.debug('failed to open XLS')

                            print('Trying to Connect and copy same data to MySQL server')
                            self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                            try:
                                db = mysql.connector.connect(
                                    host='107.180.21.18',
                                    user='grow097365',
                                    passwd='Jknm678##Tg',
                                    database='equity_property'
                                )
                                mycursor = db.cursor()
                                print(db)  # checking our connection to DB
                                command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                                print(command)

                                mycursor.execute(command)
                                myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                                print(len(myresult))
                                print(myresult)

                                if len(myresult) == 0:
                                    print('Similar homes not found, copying to database!')
                                    db = mysql.connector.connect(
                                        host='107.180.21.18',
                                        user='grow097365',
                                        passwd='Jknm678##Tg',
                                        database='equity_property'
                                    )
                                    mycursor = db.cursor()
                                    print(db)
                                    sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                    val = (self.dict_home_data['id_generated'],
                                           datetime.datetime.now(),
                                           self.dict_home_data['address'],
                                           self.short_state,
                                           self.metropolitan,
                                           self.dict_home_data['home_name'],
                                           self.dict_home_data['home_size'],
                                           self.dict_home_data['beds'],
                                           self.dict_home_data['baths'],
                                           self.dict_home_data['garage'],
                                           self.dict_home_data['priced_from'],
                                           self.dict_home_data['gallery_view_picture'],
                                           self.dict_home_data['type'],
                                           str(self.dict_home_data['id_generated_home']),
                                           self.dict_home_data['name_community'])
                                    mycursor.execute(sql, val)
                                    db.commit()
                                    time.sleep(3)
                                    print('IMPORTANT - Home floorplan data copied to mySQL')
                                else:
                                    print('Similar home found in database')
                            except:
                                print('failed to work with mySQL')

                            try:
                                driver = self.driver
                                print('trying to go back to HOMES list after data copied')
                                driver.back()
                                time.sleep(7)
                            except:
                                print('could not go back on general HOMES list')
                        print('HOMES')
                        for j in range(0, int(self.num_of_moving_homes)):
                            try:
                                driver = self.driver
                                print('entering Homes and copy the data')
                                print('For Home number {}'.format(j + 1))
                                print('Choosing quick mov in Homes')
                                driver.execute_script("window.scrollTo(0, 2050)")
                                time.sleep(3)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                                time.sleep(3)
                                print('Scrolling to Home')
                                time.sleep(3)
                                scroll = 2000 + (245 * j)
                                scroll = "window.scrollTo(0, " + str(scroll) + ")"
                                driver.execute_script(scroll)
                                print('scrolled to Homes')
                                print('trying to enter - Homes')
                            except:
                                print('could not locate floorplan home!')

                            if int(self.num_of_moving_homes) <= 1:
                                try:
                                    driver = self.driver
                                    print('if home is <= 1 , trying to find home link')
                                    print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('home link clicked')
                                    time.sleep(5)
                                    print('home entered')
                                    print('waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home LOCATED!')
                                    time.sleep(15)
                                    print('generating home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('could not locate home')
                            else:
                                try:
                                    driver = self.driver
                                    print('if Homes more then > 1')
                                    print('clicking on home link')
                                    ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                    time.sleep(5)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                    time.sleep(5)
                                    print('home link clicked')
                                    print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                    print('waiting for the home info to appear')
                                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                    print('Home LOCATED!')
                                    time.sleep(15)
                                    print('generating home id')
                                    self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                    print(type(self.dict_home_data['id_generated_home']))
                                    print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                                except:
                                    print('Home not located on path number {}'.format(j + 1))
                            try:
                                self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                                self.dict_home_data['type'] = "MIR"
                                print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))

                                time.sleep(2)
                                try:
                                    print('try copy home picture')
                                    driver = self.driver
                                    print('getting the source link of the picture')
                                    self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                                    print(self.dict_home_data['gallery_view_picture'])
                                    urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                                except:
                                    print('failed to locate pictures')
                                    self.dict_home_data['gallery_view_picture'] = 'NA'

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                    print(self.dict_home_data['home_name'])
                                except:
                                    print('home name not found')

                                try:
                                    self.dict_home_data['address'] = self.dict_community_data['address']
                                    print(self.dict_home_data['address'])
                                except:
                                    print('address not found')

                                try:
                                    self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                    print(self.dict_home_data['name_community'])
                                except:
                                    print('name community not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                    print(self.dict_home_data['home_site'])
                                except:
                                    print('home site not found')

                                self.dict_home_data['included_features_pdf'] = 'under solution'

                                try:
                                    driver = self.driver
                                    self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                    print(self.dict_home_data['availability'])
                                except:
                                    print('availability not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                    self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                    print(self.dict_home_data['priced_from'])
                                except:
                                    print('priced from not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                    print(self.dict_home_data['home_size'])
                                except:
                                    print('home size not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                    print(self.dict_home_data['stories'])
                                except:
                                    print('stories not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                    print(self.dict_home_data['beds'])
                                except:
                                    print('beds not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                                    print(self.dict_home_data['baths'])
                                except:
                                    print('baths not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                                    print(self.dict_home_data['garage'])
                                except:
                                    print('garage not found')

                                try:
                                    driver = self.driver
                                    self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                    print(self.dict_home_data['description'])
                                except:
                                    print('description not found')

                                try:
                                    driver = self.driver
                                    print('trying to copy home FloorPlan Pic scrolling')
                                    driver.execute_script("window.scrollTo(0, 1600)")
                                    time.sleep(4)
                                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                    time.sleep(3)
                                    self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                                except:
                                    print('could not locate home pics and FloorPlan Pic')
                            except:
                                print('could not locate HOME / elements')

                        print('after all homes was scanned, we going back to community')
                        try:
                            driver = self.driver
                            time.sleep(5)
                            driver.back()
                            time.sleep(5)
                            driver.back()
                            time.sleep(10)
                            print('Waiting till the page will load the community')
                        except:
                            print('could not go back on community list')
                    else:
                        print('End of community list, number of communities was {} '.format(self.num_of_communities))
                # page num
                try:
                    driver = self.driver
                    print('scrolling to next page button')
                    driver.execute_script("window.scrollTo(0, 8000)")
                    time.sleep(5)

                    for i in range(10, 0, -1):
                        try:
                            print('trying to locate next page button {}'.format(i))
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[2]/a[' + str(i) + ']').click()
                            time.sleep(5)
                            print('next button located on number {}'.format(i))
                        except:
                            print('trying another path - next page button was not located')

                    print('next page pressed')
                    time.sleep(2)
                    driver.execute_script("window.scrollTo(0, 0)")
                    print('scrolling back to top')
                    time.sleep(6)
                except:
                    print('button not located')
            print('END of work on communities > 30')
    '''

    def return_community_address_list(self):
        return self.community_address_list_full

    def return_Generated_Id_list(self):
        return self.id_random_list

    def return_list_of_homes(self):
        return self.list_of_homes


# Zillow API - api_key = "X1-ZWz1hbswvtw74b_3tnpx"
class CMA(object):
    def __init__(self, address, api_key, zip_code, xls_name):
        # all setup params
        self.address = address
        self.api_key = api_key
        self.zip_code = zip_code
        self.xls_name = xls_name

        self.dict_zillow = {
            'address': self.address,
            'api_key': self.api_key,
            'zip_code': self.zip_code,
            'zpid': '',
            'link - comparables': '',
            'link - graphs_and_data': '',
            'link - home_details': '',
            'link - map_this_home': '',
            'amount': '',
            'amount_currency': '',
            'amount_last_updated': '',
            'valuation_range_high': '',
            'valuation_range_low': '',
            'bathrooms': '',
            'bedrooms': '',
            'complete': '',
            'finished_sqft': '',
            'fips_county': '',
            'last_sold_date': '',
            'last_sold_price': '',
            'lot_size_sqft': '',
            'tax_assessment': '',
            'tax_assessment_year': '',
            'usecode': '',
            'year_built': '',

        }

    # activation zillow API and copy params to dict
    def zillow_api(self):
        try:
            locale.setlocale(locale.LC_ALL, '')
            api = zillow.ValuationApi()
            # get deep search results, also getting the zswid-ID
            data = api.GetDeepSearchResults(self.api_key, self.address, self.zip_code)
            my_dict = data.get_dict()
            # copy web page params to dictionary dict_zillow
            self.dict_zillow['zpid'] = my_dict['zpid']
            self.dict_zillow['link - comparables'] = my_dict['links']['comparables']
            self.dict_zillow['link - graphs_and_data'] = my_dict['links']['graphs_and_data']
            self.dict_zillow['link - home_details'] = my_dict['links']['home_details']
            self.dict_zillow['link - map_this_home'] = my_dict['links']['map_this_home']
            self.dict_zillow['amount'] = my_dict['zestimate']['amount']
            self.dict_zillow['amount_currency'] = my_dict['zestimate']['amount_currency']
            self.dict_zillow['amount_last_updated'] = my_dict['zestimate']['amount_last_updated']
            self.dict_zillow['valuation_range_high'] = my_dict['zestimate']['valuation_range_high']
            self.dict_zillow['valuation_range_low'] = my_dict['zestimate']['valuation_range_low']
            self.dict_zillow['bathrooms'] = my_dict['extended_data']['bathrooms']
            self.dict_zillow['bedrooms'] = my_dict['extended_data']['bedrooms']
            self.dict_zillow['complete'] = my_dict['extended_data']['complete']
            self.dict_zillow['finished_sqft'] = my_dict['extended_data']['finished_sqft']
            self.dict_zillow['fips_county'] = my_dict['extended_data']['fips_county']
            self.dict_zillow['last_sold_date'] = my_dict['extended_data']['last_sold_date']
            self.dict_zillow['last_sold_price'] = my_dict['extended_data']['last_sold_price']
            self.dict_zillow['lot_size_sqft'] = my_dict['extended_data']['lot_size_sqft']
            self.dict_zillow['tax_assessment_year'] = my_dict['extended_data']['tax_assessment_year']
            self.dict_zillow['usecode'] = my_dict['extended_data']['usecode']
            self.dict_zillow['year_built'] = my_dict['extended_data']['year_built']
            return True
        except:
            pprint('fail to get params from zillow api')
            logging.debug('fail')
            self.dict_zillow['zpid'] = 'fail to get params from zillow api'
            return False

    # printing all dicts
    def print_all(self):
        pp = pprint.PrettyPrinter(indent=4)
        pp.pprint(self.dict_zillow)

    # returning all dictionaries for future use for general list
    def return_dict_zillow(self):
        return self.dict_zillow

    # copy dict to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.address[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.address[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.address[:25]]
            # print(wb.sheetnames)
            sheet['B2'].value = self.dict_zillow['address']
            sheet['B3'].value = self.dict_zillow['api_key']
            sheet['B4'].value = self.dict_zillow['zip_code']
            sheet['B5'].value = self.dict_zillow['zpid']
            sheet['B7'].value = self.dict_zillow['link - comparables']
            sheet['B8'].value = self.dict_zillow['link - graphs_and_data']
            sheet['B9'].value = self.dict_zillow['link - home_details']
            sheet['B10'].value = self.dict_zillow['link - map_this_home']
            sheet['B12'].value = self.dict_zillow['amount']
            sheet['B13'].value = self.dict_zillow['amount_currency']
            sheet['B14'].value = self.dict_zillow['amount_last_updated']
            sheet['B15'].value = self.dict_zillow['valuation_range_high']
            sheet['B16'].value = self.dict_zillow['valuation_range_low']
            sheet['B18'].value = self.dict_zillow['bathrooms']
            sheet['B19'].value = self.dict_zillow['bedrooms']
            sheet['B20'].value = self.dict_zillow['complete']
            sheet['B21'].value = self.dict_zillow['finished_sqft']
            sheet['B22'].value = self.dict_zillow['fips_county']
            sheet['B23'].value = self.dict_zillow['last_sold_date']
            sheet['B24'].value = self.dict_zillow['last_sold_price']
            sheet['B25'].value = self.dict_zillow['lot_size_sqft']
            sheet['B26'].value = self.dict_zillow['tax_assessment']
            sheet['B27'].value = self.dict_zillow['tax_assessment_year']
            sheet['B28'].value = self.dict_zillow['usecode']
            sheet['B29'].value = self.dict_zillow['year_built']
            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Dictionaries was completed & saved in {}".format(self.xls_name))
            logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
            return True
        except:
            return False
